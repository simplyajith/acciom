from io import BytesIO

from flask import current_app
from openpyxl import load_workbook

from application.common.constants import SupportedTestClass
from application.common.createdbdetail import create_dbconnection
from application.common.splitdbdetails import split_db
from application.model.models import TestSuite, TestCase


def save_file_to_db(current_user, project_id, data, file):
    """
    Method will save suite details provided in the Excel in the tables(
    test_case,test_suite_
    Args:
        current_user(Object):  user_id of current user
        project_id(int): project id passed associated with the user
        data(data provided in the args parser):
        file(Excel file object):Excel file

    Returns: Parse Excel and save data to db.

    """
    temp_file = TestSuite(project_id=project_id, user_id=current_user,
                          excel_name=file.filename,
                          test_suite_name=data['suite_name'])
    temp_file.save_to_db()
    sheet = data['sheet_name']
    workbook = load_workbook(filename=BytesIO(file.read()))
    sheet_index = workbook.sheetnames.index(sheet)
    worksheet = workbook.worksheets[sheet_index]
    temp_data_array = [str(rows - 2) for rows in range(2, worksheet.max_row)]
    temp_data_array.append('-1')
    # row+1 to avoid index out of range error! Need to change.
    temp_test_dict = {}
    for each_col in range(0, worksheet.max_column):
        if worksheet[1][each_col].value != 'None':
            temp_test_dict.update(
                {str(worksheet[1][each_col].value): [
                    str(worksheet[rows][each_col].value)
                    for rows in range(2,
                                      worksheet.max_row)]})

    test_case_list = data['selected_case']
    for each_row in range(worksheet.max_row - 1):
        if int(temp_data_array[each_row]) in test_case_list:
            test_case_list.remove(int(temp_data_array[each_row]))
            db_list = split_db(
                temp_test_dict[current_app.config.get('DBDETAILS')][each_row])
            src_db_id = create_dbconnection(current_user,
                                            db_list['sourcedbType'].lower(),
                                            db_list['sourcedb'],
                                            db_list['sourceServer'].lower(),
                                            db_list['sourceuser'], project_id)
            target_db_id = create_dbconnection(current_user,
                                               db_list['targetdbType'].lower(),
                                               db_list['targetdb'],
                                               db_list['targetServer'].lower(),
                                               db_list['Targetuser'],
                                               project_id)
            columndata = temp_test_dict
            [current_app.config.get('COLUMNS')][each_row]
            column = {}
            if columndata == "None" or columndata.isspace():
                pass
            else:
                if ":" in columndata.strip():
                    remove_column_space = \
                        temp_test_dict[current_app.config.get('COLUMNS')][
                            each_row].replace(" ",
                                              "")
                    split_columns = remove_column_space.split(";")
                    for each_column in split_columns:
                        column_split = each_column.split(":")
                        for _ in column_split:
                            column[column_split[0]] = column_split[1]
                else:
                    remove_column_space = \
                        temp_test_dict[current_app.config.get('COLUMNS')][
                            each_row].replace(" ",
                                              "")
                    column_list = remove_column_space.split(";")
                    column = {}
                    for each_col in range(len(column_list)):
                        column[column_list[each_col]] = column_list[each_col]
            table_list = temp_test_dict[current_app.config.get('TABLES')][
                each_row].replace(" ", "").split(":")
            table = {}
            table[table_list[0]] = table_list[1]
            query = {}
            custom_queries = temp_test_dict['Custom queries'][each_row]
            if not (custom_queries == "None" or custom_queries.isspace()):
                if ";" in custom_queries:
                    query_split = custom_queries.split(";")
                    final = [each_qry.split(":") for each_qry in query_split]
                    query["sourceqry"] = final[0][1] if 'srcqry' in final[
                        0] else ""
                    query["targetqry"] = final[1][1] if 'targetqry' in final[
                        1] else ""
                else:
                    if "srcqry:" in custom_queries.lower():
                        strip_result = custom_queries.strip("srcqry:")
                        query["sourceqry"] = strip_result
                        query['targetqry'] = ""
                    elif "targetqry:" in custom_queries.lower():
                        strip_result = custom_queries.strip("targetqry:")
                        query["targetqry"] = strip_result
                        query['sourceqry'] = ""
                    else:
                        query["sourceqry"] = ""
                        query["targetqry"] = ""
            jsondict = {"column": column, "table": table, "query": query,
                        "src_db_id": src_db_id, "target_db_id": target_db_id}
            temp = TestCase(test_suite_id=temp_file.test_suite_id,
                            user_id=current_user,
                            test_case_class=SupportedTestClass().
                            get_test_class_id_by_name(
                                temp_test_dict[
                                    current_app.config.get('TESTCLASS')][
                                    each_row].lower()),
                            test_case_detail=jsondict)
            temp.save_to_db()
            data = {
                "status": True,
                "message": "success",
                "Suite": temp_file
            }
    return data
